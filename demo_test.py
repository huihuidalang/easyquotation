
# coding:utf8
"""
查询所有股票和所有可转债信息，并生成Excel文件
"""
import os
import datetime
import easyquotation
import akshare as ak
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XlImage
from kline import plot_kline


# 板块定义：名称 -> 代码前缀过滤规则
BOARD_MAP = {
    "沪A主板": lambda num, prefix: prefix == "sh" and num.startswith("60"),
    "科创板": lambda num, prefix: prefix == "sh" and num.startswith("688"),
    "深A主板": lambda num, prefix: prefix == "sz" and num.startswith("00"),
    "创业板": lambda num, prefix: prefix == "sz" and num.startswith("30"),
    "北交所": lambda num, prefix: prefix == "bj",
}

# 涨停幅度：不同板块涨跌停限制不同
LIMIT_MAP = {
    "沪A主板": 10.0,
    "科创板": 20.0,
    "深A主板": 10.0,
    "创业板": 20.0,
    "北交所": 30.0,
}

BOARD_OPTIONS = ["全部"] + list(BOARD_MAP.keys())

LIMIT_OPTIONS = ["涨停", "跌停"]


def _match_board(code, board_name):
    """判断代码是否属于指定板块"""
    prefix = code[:2]
    num = code[2:]
    if board_name == "全部":
        return any(rule(num, prefix) for rule in BOARD_MAP.values())
    return BOARD_MAP[board_name](num, prefix)


def _get_board_name(code):
    """根据代码返回所属板块名称"""
    prefix = code[:2]
    num = code[2:]
    for name, rule in BOARD_MAP.items():
        if rule(num, prefix):
            return name
    return "其他"



def _fetch_history_by_date(date_str):
    """通过腾讯接口获取指定日期的全市场行情数据
    date_str: 日期字符串，如 "20250110" 或 "2025-01-10"
    返回: dict {code: {name, open, close, now, high, low, turnover, volume, date}}

    优化思路：
    1. 使用新浪接口获取股票代码和名称列表（1次批量请求）
    2. 如果目标日期是今天，直接使用新浪实时行情数据，无需逐只查询
    3. 历史日期使用腾讯历史K线接口，多线程并发加速
    """
    import json
    import re
    import requests
    import datetime as _dt
    from concurrent.futures import ThreadPoolExecutor, as_completed

    d = date_str.replace("-", "")
    formatted_date = f"{d[:4]}-{d[4:6]}-{d[6:8]}"

    try:
        target_dt = _dt.datetime.strptime(d, "%Y%m%d")
    except ValueError:
        print(f"日期格式错误: {date_str}")
        return {}

    # 判断目标日期是否为今天
    today_str = _dt.datetime.now().strftime("%Y%m%d")
    if d == today_str:
        print(f"目标日期为今天，直接使用新浪实时行情数据...")
        quotation = easyquotation.use("sina")
        return quotation.market_snapshot(prefix=True)

    # 历史日期：获取股票列表 + 逐只腾讯K线查询
    print("正在获取股票代码列表...")
    quotation = easyquotation.use("sina")
    snapshot = quotation.market_snapshot(prefix=True)
    codes = list(snapshot.keys())
    print(f"共 {len(codes)} 只股票，开始获取 {formatted_date} 历史行情...")

    def fetch_one(code):
        info = snapshot.get(code, {})
        name = info.get("name", "")

        # 腾讯K线接口symbol格式: sh600000, sz000001
        # 北交所股票腾讯不支持，跳过
        if code.startswith("bj"):
            return None

        symbol = code

        # 使用不复权数据计算涨跌幅（前复权会调整历史价格，导致涨跌幅计算错误）
        url = "https://web.ifzq.gtimg.cn/appstock/app/fqkline/get"
        params = {
            "_var": "kline_day",
            "param": f"{symbol},day,,,15,",
        }
        try:
            r = requests.get(url, params=params, timeout=5)
            text = r.text
            # 解析JS变量赋值: kline_day={...}
            match = re.search(r'kline_day=(\{.*\})', text, re.DOTALL)
            if not match:
                return None
            data = json.loads(match.group(1))
        except Exception:
            return None

        if data.get("code") != 0:
            return None

        # 提取K线数据（不复权）
        stock_data = data.get("data", {}).get(symbol, {})
        klines = stock_data.get("day", [])
        if not klines:
            return None

        # 找目标日期的行和前一日收盘价
        # 腾讯返回格式: [日期, 开盘, 收盘, 最高, 最低, 成交量]
        target_row = None
        prev_row = None
        for i, item in enumerate(klines):
            row_date = str(item[0]).replace("-", "")
            if row_date == d:
                target_row = item
                # 取紧邻的前一条K线作为昨收
                if i > 0:
                    prev_row = klines[i - 1]
                break
            # 保留最后一条早于目标日期的记录
            if row_date < d:
                prev_row = item

        # 计算昨收价（不复权）
        prev_close = 0
        if prev_row is not None:
            try:
                prev_close = float(prev_row[2])
            except (ValueError, TypeError, IndexError):
                prev_close = 0

        if target_row is None:
            return None

        try:
            open_price = float(target_row[1])
            close_price = float(target_row[2])
            high_price = float(target_row[3])
            low_price = float(target_row[4])
            volume_amt = float(target_row[5]) if len(target_row) > 5 else 0
        except (ValueError, TypeError, IndexError):
            return None

        return (code, {
            "name": name,
            "open": open_price,
            "close": prev_close,
            "now": close_price,
            "high": high_price,
            "low": low_price,
            "turnover": 0,  # 腾讯K线不提供换手率
            "volume": volume_amt,
            "date": formatted_date,
            "time": "",
        })

    result = {}
    total = len(codes)
    done_count = 0
    fail_count = 0
    with ThreadPoolExecutor(max_workers=20) as executor:
        futures = {executor.submit(fetch_one, code): code for code in codes}
        for future in as_completed(futures):
            done_count += 1
            if done_count % 500 == 0:
                print(f"  进度: {done_count}/{total}，已获取 {len(result)} 只...")
            try:
                res = future.result()
                if res:
                    code, item = res
                    result[code] = item
                else:
                    fail_count += 1
            except Exception:
                fail_count += 1

    print(f"获取到 {len(result)} 只股票（{formatted_date}），失败 {fail_count} 只")
    return result














def get_all_stocks_data(board="全部", change_pct_min=None, change_pct_max=None, limit=None, date_str=None):
    """获取A股行情数据
    board: 板块选项，可选 "全部"/"沪A主板"/"科创板"/"深A主板"/"创业板"/"北交所"
    change_pct_min: 涨跌幅最小值(%)，None表示不限制
    change_pct_max: 涨跌幅最大值(%)，None表示不限制
    limit: 涨跌停选项，可选 "涨停"/"跌停"/None
    date_str: 指定日期，如 "20250110" 或 "2025-01-10"，None则获取实时行情
    """
    # 涨停/跌停模式：按各板块涨跌停幅度筛选
    is_limit_up = (limit == "涨停")
    is_limit_down = (limit == "跌停")

    filter_parts = [board]
    if is_limit_up:
        filter_parts.append("涨停")
    elif is_limit_down:
        filter_parts.append("跌停")
    elif change_pct_min is not None or change_pct_max is not None:
        filter_parts.append(f"涨跌幅{change_pct_min}%~{change_pct_max}%")
    if date_str:
        filter_parts.append(f"日期:{date_str}")
    filter_desc = "，".join(filter_parts)
    print(f"正在获取{'历史' if date_str else '实时'}行情（{filter_desc}），请稍候...")

    # 获取行情数据：指定日期用历史接口，否则用实时接口
    if date_str:
        data = _fetch_history_by_date(date_str)
        if not data:
            print(f"警告：未获取到 {date_str} 的历史行情数据")
            return pd.DataFrame()
    else:
        quotation = easyquotation.use("sina")
        data = quotation.market_snapshot(prefix=True)

    rows = []
    for code, info in data.items():
        # 按板块过滤
        if not _match_board(code, board):
            continue

        name = info.get("name", "")
        now = info.get("now", 0)
        # 过滤现价为0的无效数据（停牌或无行情）
        if now <= 0:
            continue
        close = info.get("close", 0)
        if close <= 0:
            # 昨收为0则无法计算涨跌幅，跳过
            continue
        change_pct = ((now / close) - 1) * 100

        # 涨停/跌停过滤：使用涨停价/跌停价精确判断
        board_name = _get_board_name(code)
        limit_pct = LIMIT_MAP.get(board_name, 10.0)
        # 涨停价 = 昨收 × (1 + 涨停幅度/100)，四舍五入到2位小数
        limit_up_price = round(close * (1 + limit_pct / 100), 2)
        # 跌停价 = 昨收 × (1 - 涨停幅度/100)，四舍五入到2位小数
        limit_down_price = round(close * (1 - limit_pct / 100), 2)
        if is_limit_up:
            # 涨停：现价 >= 涨停价（允许0.01误差，因四舍五入）
            if now < limit_up_price - 0.01:
                continue
        elif is_limit_down:
            # 跌停：现价 <= 跌停价（允许0.01误差）
            if now > limit_down_price + 0.01:
                continue
        else:
            # 普通涨跌幅过滤
            if change_pct_min is not None and change_pct < change_pct_min:
                continue
            if change_pct_max is not None and change_pct > change_pct_max:
                continue
        rows.append({
            "板块": _get_board_name(code),
            "代码": code,
            "名称": name,
            "现价": now,
            "涨跌幅(%)": round(change_pct, 2),
            "开盘价": info.get("open", 0),
            "最高价": info.get("high", 0),
            "最低价": info.get("low", 0),
            "昨收": close,
            "成交量(手)": info.get("turnover", 0),
            "成交额": info.get("volume", 0),
            "日期": info.get("date", ""),
            "时间": info.get("time", ""),
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(by="代码").reset_index(drop=True)
    print(f"获取到 {len(df)} 只股票（{filter_desc}）")
    return df


def get_all_convertible_bonds_data():
    """获取全部在市可转债实时行情数据（已过滤退市）"""
    from datetime import datetime as dt

    # 第一步：从同花顺获取可转债列表，按到期时间过滤掉已退市的
    print("正在获取可转债列表（同花顺），请稍候...")
    try:
        cb_list_df = ak.bond_zh_cov_info_ths()
        # 只保留有上市日期的
        cb_list_df = cb_list_df[cb_list_df["上市日期"].notna()]
        # 按到期时间过滤：只保留未到期的（在市可转债）
        today = dt.now().date()
        cb_list_df["到期日期"] = cb_list_df["到期时间"].apply(
            lambda x: x.date() if hasattr(x, "date") else x
        )
        active_df = cb_list_df[cb_list_df["到期日期"] >= today]
        cb_codes = active_df["债券代码"].tolist()
        print(f"同花顺可转债列表: 总计 {len(cb_list_df)} 只，未到期(在市) {len(active_df)} 只，已到期(退市) {len(cb_list_df) - len(active_df)} 只")
    except Exception as e:
        print(f"同花顺接口失败: {e}，使用默认代码范围")
        cb_codes = []
        for i in range(110000, 114000):
            cb_codes.append(str(i))
        for i in range(123000, 129000):
            cb_codes.append(str(i))

    # 第二步：用新浪查询实时行情
    print("正在查询可转债实时行情，请稍候...")
    quotation = easyquotation.use("sina")
    data = quotation.real(cb_codes, prefix=True)

    rows = []
    for code, info in data.items():
        name = info.get("name", "")
        now = info.get("now", 0)
        turnover = info.get("turnover", 0)
        open_price = info.get("open", 0)
        if name and name != "N/A" and now > 0:
            # 过滤已退市转债（强赎退市）：成交量为0且开盘价为0说明已无交易
            if turnover == 0 and open_price == 0:
                continue
            close = info.get("close", 0)
            change_pct = ((now / close) - 1) * 100 if close else 0
            rows.append({
                "代码": code,
                "名称": name,
                "现价": now,
                "涨跌幅(%)": round(change_pct, 2),
                "开盘价": open_price,
                "最高价": info.get("high", 0),
                "最低价": info.get("low", 0),
                "昨收": close,
                "成交量": turnover,
            })

    df = pd.DataFrame(rows)
    df = df.sort_values(by="代码").reset_index(drop=True)
    print(f"获取到 {len(df)} 只可转债（已过滤到期退市+强赎退市）")
    return df


def _style_worksheet(ws):
    """为worksheet设置样式"""
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 设置表头样式
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 设置数据区域样式
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 涨跌幅列标红/标绿
            header_val = ws.cell(row=1, column=col_idx).value
            if header_val and "涨跌幅" in str(header_val):
                try:
                    val = float(cell.value)
                    if val > 0:
                        cell.font = Font(color="FF0000")  # 红色涨
                    elif val < 0:
                        cell.font = Font(color="008000")  # 绿色跌
                except (ValueError, TypeError):
                    pass

    # 自动调整列宽
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        for row_idx in range(1, min(ws.max_row + 1, 100)):  # 采样前100行
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                val_str = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val_str)
                max_length = max(max_length, length)
        ws.column_dimensions[col_letter].width = max_length + 4

    # 冻结首行
    ws.freeze_panes = "A2"


def save_to_excel(stock_df, cb_df, output_dir=None, kline_dir=None,
                   show_kline=False, kline_days=120, kline_indicators=None,
                   board="全部", limit=None, date_str=None):
    """将股票和可转债数据保存到Excel文件，按板块分sheet
    :param stock_df: 股票数据DataFrame
    :param cb_df: 可转债数据DataFrame
    :param output_dir: Excel输出目录
    :param kline_dir: K线图输出目录（None则使用output_dir下的kline子目录）
    :param show_kline: 是否为每只股票生成K线图并嵌入Excel
    :param kline_days: K线天数
    :param kline_indicators: 技术指标列表
    :param board: 板块选项
    :param limit: 涨跌停选项，"涨停"/"跌停"/None
    :param date_str: 指定日期字符串，如 "20250110"，None则使用当前时间
    """
    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(__file__))
    if kline_dir is None:
        kline_dir = os.path.join(output_dir, "kline")
    if kline_indicators is None:
        kline_indicators = ["macd", "kdj", "rsi", "boll", "dmi", "expma"]

    if date_str:
        filename = os.path.join(output_dir, f"行情数据_{date_str}.xlsx")
    else:
        now = datetime.datetime.now()
        filename = os.path.join(output_dir, f"行情数据_{now.strftime('%Y%m%d_%H%M%S')}.xlsx")

    # ========= 生成K线图 =========
    kline_images = {}  # code -> image_path
    if show_kline and len(stock_df) > 0:
        os.makedirs(kline_dir, exist_ok=True)
        total = len(stock_df)
        print(f"\n正在生成K线图（共{total}只，指标: {','.join(kline_indicators)}）...")
        for idx, row in stock_df.iterrows():
            code = row["代码"]
            name = row.get("名称", "")
            print(f"  [{idx+1}/{total}] {name}({code})...", end=" ")
            try:
                img_path = plot_kline(
                    code, days=kline_days, output_dir=kline_dir,
                    show_indicators=kline_indicators,
                )
                if img_path:
                    kline_images[code] = img_path
                    print("OK")
                else:
                    print("无数据")
            except Exception as e:
                print(f"失败: {e}")

    print(f"\n正在生成Excel文件: {filename}")

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # 涨停/跌停模式：按板块分别输出sheet + 一个汇总sheet
        if stock_df.empty:
            # 空数据时写入空sheet
            pd.DataFrame().to_excel(writer, sheet_name="股票数据(空)", index=False)
        elif limit in ("涨停", "跌停"):
            # 先写汇总sheet
            stock_df.to_excel(writer, sheet_name=limit, index=False)
            # 再按板块分别写入
            for board_name in BOARD_MAP.keys():
                board_df = stock_df[stock_df["板块"] == board_name]
                if len(board_df) > 0:
                    sheet_name = f"{board_name}{limit}"
                    board_df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # 按板块分别写入sheet
            for board_name in list(BOARD_MAP.keys()) + ["全部"]:
                if board_name == "全部":
                    board_df = stock_df
                else:
                    board_df = stock_df[stock_df["板块"] == board_name]
                if len(board_df) > 0:
                    sheet_name = board_name
                    board_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # 写入可转债数据
        cb_df.to_excel(writer, sheet_name="全部可转债", index=False)

        # 获取workbook对象进行样式设置
        wb = writer.book
        for ws in wb.worksheets:
            _style_worksheet(ws)

    # ========= 嵌入K线图到Excel =========
    if kline_images:
        print("\n正在嵌入K线图到Excel...")
        from openpyxl import load_workbook
        wb = load_workbook(filename)

        for ws in wb.worksheets:
            if ws.title == "全部可转债":
                continue
            # 找到"代码"列的位置
            code_col = None
            for col_idx in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col_idx).value == "代码":
                    code_col = col_idx
                    break
            if code_col is None:
                continue

            # 在最后一列后面添加"K线图"列标题
            img_col = ws.max_column + 1
            ws.cell(row=1, column=img_col, value="K线图")
            # 设置表头样式
            header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=1, column=img_col).font = header_font
            ws.cell(row=1, column=img_col).fill = header_fill
            ws.cell(row=1, column=img_col).alignment = header_alignment

            # 为每行股票插入K线图
            for row_idx in range(2, ws.max_row + 1):
                code = str(ws.cell(row=row_idx, column=code_col).value)
                if code in kline_images:
                    img_path = kline_images[code]
                    try:
                        img = XlImage(img_path)
                        # 缩小图片尺寸以适合Excel
                        orig_width = img.width
                        orig_height = img.height
                        target_width = 480
                        if orig_width > 0:
                            scale = target_width / orig_width
                            img.width = target_width
                            img.height = int(orig_height * scale)
                        # 设置行高以容纳图片
                        ws.row_dimensions[row_idx].height = 160
                        # 插入图片到对应单元格
                        cell_ref = ws.cell(row=row_idx, column=img_col).coordinate
                        ws.add_image(img, cell_ref)
                    except Exception as e:
                        ws.cell(row=row_idx, column=img_col, value="图加载失败")

        # 设置K线图列宽
        for ws in wb.worksheets:
            if ws.title == "全部可转债":
                continue
            img_col = ws.max_column
            if ws.cell(row=1, column=img_col).value == "K线图":
                ws.column_dimensions[ws.cell(row=1, column=img_col).column_letter].width = 55

        wb.save(filename)
        print(f"K线图嵌入完成: {len(kline_images)} 只")

    print(f"Excel文件已生成: {filename}")
    # 打印各sheet数量
    if stock_df.empty:
        print("  - 股票数据为空")
    else:
        for board_name in list(BOARD_MAP.keys()) + ["全部"]:
            if board_name == "全部":
                count = len(stock_df)
            else:
                count = len(stock_df[stock_df["板块"] == board_name])
            if count > 0:
                print(f"  - {board_name}: {count} 只")
    print(f"  - 全部可转债: {len(cb_df)} 只")
    return filename


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="全部股票与可转债信息查询")
    parser.add_argument("--board", type=str, default="全部", choices=BOARD_OPTIONS,
                        help="板块选择：全部/沪A主板/科创板/深A主板/创业板/北交所")
    parser.add_argument("--limit", type=str, default=None, choices=LIMIT_OPTIONS,
                        help="涨跌停筛选：涨停/跌停")
    parser.add_argument("--date", type=str, default=None,
                        help="指定日期，格式: YYYYMMDD，如 20250110，默认使用当前时间")
    parser.add_argument("--change_pct_min", type=float, default=None,
                        help="涨跌幅最小值(%%)，例如: -5 表示跌5%%以上")
    parser.add_argument("--change_pct_max", type=float, default=None,
                        help="涨跌幅最大值(%%)，例如: 5 表示涨5%%以内")
    parser.add_argument("--kline", action="store_true", default=False,
                        help="是否生成K线图并嵌入Excel（默认不生成）")
    parser.add_argument("--kline_days", type=int, default=120,
                        help="K线天数，默认120")
    parser.add_argument("--indicators", type=str, default="macd,kdj,rsi,boll,dmi,expma",
                        help="技术指标，逗号分隔，可选: macd,kdj,rsi,boll,dmi,expma")
    args = parser.parse_args()

    kline_indicators = [x.strip() for x in args.indicators.split(",")]

    print("\n全部股票与可转债信息查询\n")
    stock_df = get_all_stocks_data(board=args.board, change_pct_min=args.change_pct_min, change_pct_max=args.change_pct_max, limit=args.limit, date_str=args.date)
    cb_df = get_all_convertible_bonds_data()
    save_to_excel(stock_df, cb_df,
                  show_kline=args.kline,
                  kline_days=args.kline_days,
                  kline_indicators=kline_indicators,
                  board=args.board,
                  limit=args.limit,
                  date_str=args.date)
    print("\n查询完成！")
